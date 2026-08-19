"""人员清单 ↔ 安全质量奖核算数据 的解析与对账。

四类判定（按需求原文）：
  1. 新入职员工        —— 清单"生产部"有、核算"一线人员"无，且入职时间在参照日期一个月内，
                          或出现在"人员变动说明"且离职时间为空
  2. 待定需填入人员(A) —— 清单有、核算无，但不满足上面的新入职条件
  3. 离职人员          —— 核算"一线人员"有、清单"生产部"无，且在"离职人员&调出人员"，
                          或在"人员变动说明"且有离职时间
  4. 待定需填入人员(B) —— 核算有、清单无，但不满足上面的离职条件
"""

from __future__ import annotations

import datetime as _dt
import io
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field

from .normalize import (
    clean_text,
    fmt_date,
    is_blank,
    key_label,
    months_before,
    norm_eid,
    norm_name,
    parse_date,
    date_from_filename,
)

# 需求指定的目标职务；"助理工程师" 在核算表里写作 "助工"
TARGET_DUTIES = ("助工", "助理工程师", "操作工", "工程师", "班长")
DUTY_ALIAS = {"助理工程师": "助工"}

CATEGORY_NEW = "新入职员工"
CATEGORY_PENDING_ADD = "待定需填入人员（清单有·核算无）"
CATEGORY_LEFT = "离职人员"
CATEGORY_PENDING_DEL = "待定需填入人员（核算有·清单无）"

CATEGORIES = (CATEGORY_NEW, CATEGORY_PENDING_ADD, CATEGORY_LEFT, CATEGORY_PENDING_DEL)
ADD_CATEGORIES = (CATEGORY_NEW, CATEGORY_PENDING_ADD)
REMOVE_CATEGORIES = (CATEGORY_LEFT, CATEGORY_PENDING_DEL)

SHEET_PRODUCTION = "生产部"
SHEET_EQUIPMENT = "生产设备部"
SHEET_DEPARTURE = "离职人员&调出人员"
SHEET_CHANGES = "人员变动说明"
SHEET_FRONTLINE = "一线人员"
SHEET_OTHERS = "副主任&工艺组长及其他"

NEW_BLOCK_SENTINEL = "__新增车间__"


def canon_duty(value) -> str:
    text = clean_text(value)
    return DUTY_ALIAS.get(text, text)


def is_target_duty(value) -> bool:
    return canon_duty(value) in {DUTY_ALIAS.get(d, d) for d in TARGET_DUTIES}


# --------------------------------------------------------------------------- #
# 通用定位工具
# --------------------------------------------------------------------------- #


def find_sheet(workbook, *candidates: str):
    """按名字精确/模糊定位工作表。"""
    titles = {clean_text(ws.title): ws for ws in workbook.worksheets}
    for candidate in candidates:
        target = clean_text(candidate)
        if target in titles:
            return titles[target]
    for candidate in candidates:
        target = clean_text(candidate)
        for title, sheet in titles.items():
            if target and (target in title or title in target):
                return sheet
    return None


def find_col(header: list, *candidates: str) -> int | None:
    """在表头里找列（0 基）。先精确，再前缀/包含。"""
    cleaned = [clean_text(cell) for cell in header]
    for candidate in candidates:
        target = clean_text(candidate)
        for index, cell in enumerate(cleaned):
            if cell == target:
                return index
    for candidate in candidates:
        target = clean_text(candidate)
        if not target:
            continue
        for index, cell in enumerate(cleaned):
            if cell and (cell.startswith(target) or target in cell):
                return index
    return None


def _get(row: tuple, index: int | None):
    if index is None or index >= len(row):
        return None
    return row[index]


# --------------------------------------------------------------------------- #
# 数据模型
# --------------------------------------------------------------------------- #


@dataclass
class Person:
    name: str
    eid: str
    duty: str
    duty_raw: str
    group: str
    hire_date: _dt.date | None
    remark: str = ""
    leave_date: _dt.date | None = None
    leave_raw: str = ""
    row: int = 0
    workshop: str = ""
    post: str = ""  # 清单的「岗位」列
    title: str = ""  # 清单的「职位」列
    is_intern: bool = False
    intern_source: str = ""  # 「职位」或「备注」

    @property
    def key(self) -> tuple[str, str]:
        return (self.name, self.eid)


@dataclass
class SheetLayout:
    """一张"按车间分组、车间内按职务分段"的人员子表的版式信息。"""

    name: str
    columns: dict[str, str] = field(default_factory=dict)
    first_data_row: int = 2
    last_data_row: int = 2
    merged_workshop: bool = False
    people: dict[tuple[str, str], Person] = field(default_factory=dict)
    blocks: list[tuple[str, int, int]] = field(default_factory=list)
    duty_anchors: dict[tuple[str, str], int] = field(default_factory=dict)
    duty_order: dict[str, list[str]] = field(default_factory=dict)
    # 规范化后的车间名 -> 单元格里的原始文本。匹配用规范化名，写回必须用原文，
    # 否则「生产技术转移组（多肽）」会被写成半角括号，Excel 里就成了两个不同的车间。
    raw_names: dict[str, str] = field(default_factory=dict)

    @property
    def workshops(self) -> list[str]:
        return [block[0] for block in self.blocks]

    def raw_workshop(self, workshop: str) -> str:
        return self.raw_names.get(workshop, workshop)

    def block_of(self, workshop: str) -> tuple[str, int, int] | None:
        for block in self.blocks:
            if block[0] == workshop:
                return block
        return None

    def anchor_for(self, workshop: str, duty: str) -> tuple[int, str]:
        """返回 (插入锚点行, 定位精度)：职务段末 / 车间块末 / 全表末。"""
        anchor = self.duty_anchors.get((workshop, canon_duty(duty)))
        if anchor is not None:
            return anchor, "职务"
        block = self.block_of(workshop)
        if block is not None:
            return block[2], "车间"
        return self.last_data_row, "表尾"


@dataclass
class RosterFile:
    """TJ4 人员清单。"""

    file_name: str
    ref_date: _dt.date | None
    production: dict[tuple[str, str], Person] = field(default_factory=dict)
    production_all: dict[tuple[str, str], Person] = field(default_factory=dict)
    equipment: dict[tuple[str, str], Person] = field(default_factory=dict)
    departures: dict[tuple[str, str], Person] = field(default_factory=dict)
    changes: dict[tuple[str, str], dict] = field(default_factory=dict)
    duty_field: str = "职位"
    notes: list[str] = field(default_factory=list)


@dataclass
class BonusFile:
    """安全质量奖核算数据。"""

    file_name: str
    frontline: dict[tuple[str, str], Person] = field(default_factory=dict)
    frontline_order: list[tuple[str, str]] = field(default_factory=list)
    blocks: list[tuple[str, int, int]] = field(default_factory=list)
    others: dict[tuple[str, str], Person] = field(default_factory=dict)
    others_layout: SheetLayout | None = None
    first_data_row: int = 3
    last_data_row: int = 3
    columns: dict[str, str] = field(default_factory=dict)
    notes: list[str] = field(default_factory=list)
    duty_anchors: dict[tuple[str, str], int] = field(default_factory=dict)
    duty_order: dict[str, list[str]] = field(default_factory=dict)

    def block_of(self, workshop: str) -> tuple[str, int, int] | None:
        for block in self.blocks:
            if block[0] == workshop:
                return block
        return None

    @property
    def workshops(self) -> list[str]:
        return [block[0] for block in self.blocks]

    def anchor_for(self, workshop: str, duty: str) -> tuple[int | None, bool]:
        """返回 (插入锚点行, 是否按职务精确定位)。

        锚点是该车间块里这个职务最下面的一行；该车间没有这个职务时退回块末尾。
        """
        anchor = self.duty_anchors.get((workshop, canon_duty(duty)))
        if anchor is not None:
            return anchor, True
        block = self.block_of(workshop)
        return (block[2] if block else None), False


# --------------------------------------------------------------------------- #
# 人员清单解析
# --------------------------------------------------------------------------- #


def parse_roster(
    data: bytes,
    file_name: str,
    duty_field: str = "职位",
    include_interns: bool = False,
) -> RosterFile:
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    result = RosterFile(file_name=file_name, ref_date=date_from_filename(file_name), duty_field=duty_field)
    try:
        production = find_sheet(workbook, SHEET_PRODUCTION)
        if production is None:
            raise ValueError(f"{file_name} 里找不到「{SHEET_PRODUCTION}」子表")
        result.production_all, result.production = _parse_staff_sheet(
            production, duty_field, include_interns
        )

        equipment = find_sheet(workbook, SHEET_EQUIPMENT)
        if equipment is not None:
            result.equipment = _parse_staff_sheet(equipment, duty_field, include_interns)[0]

        departure = find_sheet(workbook, SHEET_DEPARTURE)
        if departure is None:
            result.notes.append(f"未找到「{SHEET_DEPARTURE}」子表，离职判定将只依赖「{SHEET_CHANGES}」")
        else:
            result.departures = _parse_departure_sheet(departure)

        changes = find_sheet(workbook, SHEET_CHANGES)
        if changes is None:
            result.notes.append(f"未找到「{SHEET_CHANGES}」子表，新入职判定将只依赖入职时间")
        else:
            result.changes = _parse_changes_sheet(changes)
    finally:
        workbook.close()
    if result.ref_date is None:
        result.notes.append("文件名里未识别出参照日期，请在界面上手工指定")
    return result


def _staff_header(sheet):
    header = [cell.value for cell in sheet[1]]
    return header, {
        "name": find_col(header, "姓名"),
        "eid": find_col(header, "员工编号", "员工号", "工号"),
        "hire": find_col(header, "入职时间", "入职日期"),
        "post": find_col(header, "岗位"),
        "title": find_col(header, "职位", "职务"),
        "group": find_col(header, "目前分组", "目前二级分组", "分组"),
        "remark": find_col(header, "备注"),
        "leave": find_col(header, "离职时间", "离职日期"),
    }


INTERN_TITLE = "实习生"
# 有的实习生「职位」和「岗位」都写的是实际岗位，只有备注里写着"校招实习生"
INTERN_REMARK_RE = re.compile(r"实习")


def _parse_staff_sheet(sheet, duty_field: str, include_interns: bool = False):
    """返回 (全部人员, 目标职务人员)。

    实习生有两种写法：
      * 「职位」直接写「实习生」（此时「岗位」才是实际岗位）——按职位过滤会天然排除，
        ``include_interns`` 打开后改用岗位把他们捞回来；
      * 「职位」「岗位」都写实际岗位，只有**备注**里写着"校招实习生"——这些人本来就在
        目标名单里，只是需要标出来，所以不受 ``include_interns`` 影响。
    """
    _, cols = _staff_header(sheet)
    everyone: dict[tuple[str, str], Person] = {}
    targets: dict[tuple[str, str], Person] = {}
    for offset, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        name = norm_name(_get(row, cols["name"]))
        if not name:
            continue
        post = clean_text(_get(row, cols["post"]))
        title = clean_text(_get(row, cols["title"]))
        remark = clean_text(_get(row, cols["remark"]))
        by_title = title == INTERN_TITLE
        by_remark = bool(INTERN_REMARK_RE.search(remark))
        duty_raw = title if duty_field == "职位" else post
        if by_title and include_interns and not is_target_duty(duty_raw):
            duty_raw = post
        person = Person(
            name=name,
            eid=norm_eid(_get(row, cols["eid"])),
            duty=canon_duty(duty_raw),
            duty_raw=duty_raw,
            group=clean_text(_get(row, cols["group"])),
            hire_date=parse_date(_get(row, cols["hire"])),
            remark=remark,
            row=offset,
            post=post,
            title=title,
            is_intern=by_title or by_remark,
            intern_source="职位" if by_title else ("备注" if by_remark else ""),
        )
        everyone[person.key] = person
        # 只有"职位写实习生"的人受开关控制；备注型实习生的职位本来就是目标职务
        if is_target_duty(duty_raw) and (include_interns or not by_title):
            targets[person.key] = person
    return everyone, targets


def _parse_departure_sheet(sheet) -> dict[tuple[str, str], Person]:
    _, cols = _staff_header(sheet)
    out: dict[tuple[str, str], Person] = {}
    for offset, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        name = norm_name(_get(row, cols["name"]))
        if not name:
            continue
        leave_raw = clean_text(_get(row, cols["leave"]))
        person = Person(
            name=name,
            eid=norm_eid(_get(row, cols["eid"])),
            duty=canon_duty(_get(row, cols["title"]) or _get(row, cols["post"])),
            duty_raw=clean_text(_get(row, cols["title"]) or _get(row, cols["post"])),
            group=clean_text(_get(row, cols["group"])),
            hire_date=parse_date(_get(row, cols["hire"])),
            remark=clean_text(_get(row, cols["remark"])),
            leave_date=parse_date(_get(row, cols["leave"])),
            leave_raw=leave_raw,
            row=offset,
        )
        out[person.key] = person
    return out


def _parse_changes_sheet(sheet) -> dict[tuple[str, str], dict]:
    """「人员变动说明」的表头不在第 1 行（第 1 行是合并标题），需要探测。"""
    header_row = 1
    for index in range(1, min(sheet.max_row, 8) + 1):
        values = [clean_text(cell.value) for cell in sheet[index]]
        if "姓名" in values:
            header_row = index
            break
    header = [cell.value for cell in sheet[header_row]]
    cols = {
        "kind": find_col(header, "分类（入职/离职/调动）", "分类"),
        "group": find_col(header, "分组"),
        "name": find_col(header, "姓名"),
        "eid": find_col(header, "员工号", "员工编号", "工号"),
        "post": find_col(header, "岗位", "职位"),
        "hire": find_col(header, "入职时间", "入职日期"),
        "leave": find_col(header, "离职时间", "离职日期"),
        "remark": find_col(header, "备注"),
    }
    out: dict[tuple[str, str], dict] = {}
    for offset, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        name = norm_name(_get(row, cols["name"]))
        if not name:
            continue
        leave_value = _get(row, cols["leave"])
        out[(name, norm_eid(_get(row, cols["eid"])))] = {
            "kind": clean_text(_get(row, cols["kind"])),
            "group": clean_text(_get(row, cols["group"])),
            "post": clean_text(_get(row, cols["post"])),
            "hire_date": parse_date(_get(row, cols["hire"])),
            "leave_date": parse_date(leave_value),
            "leave_blank": is_blank(leave_value),
            "leave_raw": clean_text(leave_value),
            "remark": clean_text(_get(row, cols["remark"])),
            "row": offset,
        }
    return out


# --------------------------------------------------------------------------- #
# 核算数据解析
# --------------------------------------------------------------------------- #


def parse_bonus(data: bytes, file_name: str) -> BonusFile:
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    result = BonusFile(file_name=file_name)
    try:
        frontline = find_sheet(workbook, SHEET_FRONTLINE)
        if frontline is None:
            raise ValueError(f"{file_name} 里找不到「{SHEET_FRONTLINE}」子表")
        _parse_frontline(frontline, result)
        others = find_sheet(workbook, SHEET_OTHERS)
        if others is None:
            result.notes.append(f"未找到「{SHEET_OTHERS}」子表，无法排除已在该表的人员")
        else:
            result.others_layout = _parse_others(others)
            result.others = result.others_layout.people
    finally:
        workbook.close()
    return result


def _frontline_header(sheet) -> tuple[int, dict[str, str]]:
    """返回 (数据起始行, {字段: 列字母})。表头可能占两行。"""
    from openpyxl.utils import get_column_letter

    best_row, best_cols, best_hits = 1, {}, -1
    for row_index in range(1, min(sheet.max_row, 4) + 1):
        header = [cell.value for cell in sheet[row_index]]
        cols = {
            "workshop": find_col(header, "车间"),
            "duty": find_col(header, "职务", "岗位"),
            "name": find_col(header, "姓名"),
            "eid": find_col(header, "员工编号", "员工号", "工号"),
            "hire": find_col(header, "入职日期", "入职时间"),
        }
        hits = sum(1 for value in cols.values() if value is not None)
        if hits > best_hits:
            best_row, best_cols, best_hits = row_index, cols, hits
    # 表头行可能是合并的两行，数据从表头下一行开始；再向下跳过空姓名行
    name_index = best_cols.get("name")
    data_row = best_row + 1
    while data_row <= sheet.max_row:
        value = sheet.cell(data_row, (name_index or 2) + 1).value
        if norm_name(value):
            break
        data_row += 1
    letters = {
        key: get_column_letter(index + 1) for key, index in best_cols.items() if index is not None
    }
    return data_row, letters


def _parse_frontline(sheet, result: BonusFile) -> None:
    from openpyxl.utils import column_index_from_string

    data_row, letters = _frontline_header(sheet)
    result.first_data_row = data_row
    result.columns = letters
    required = ("workshop", "duty", "name", "eid", "hire")
    missing = [key for key in required if key not in letters]
    if missing:
        raise ValueError(f"「{SHEET_FRONTLINE}」缺少列：{missing}")

    ws_col = column_index_from_string(letters["workshop"])
    duty_col = column_index_from_string(letters["duty"])
    name_col = column_index_from_string(letters["name"])
    eid_col = column_index_from_string(letters["eid"])
    hire_col = column_index_from_string(letters["hire"])

    # 车间列纵向合并：先建立"行 -> 车间"映射
    merged: dict[int, str] = {}
    for merge in sheet.merged_cells.ranges:
        if merge.min_col != ws_col or merge.max_col != ws_col:
            continue
        value = clean_text(sheet.cell(merge.min_row, ws_col).value)
        if not value:
            continue
        for row in range(merge.min_row, merge.max_row + 1):
            merged[row] = value

    order: list[str] = []
    spans: dict[str, list[int]] = {}
    current = ""
    for row in range(data_row, sheet.max_row + 1):
        label = merged.get(row) or clean_text(sheet.cell(row, ws_col).value)
        if label:
            current = label
        name = norm_name(sheet.cell(row, name_col).value)
        if not name:
            continue
        eid = norm_eid(sheet.cell(row, eid_col).value)
        person = Person(
            name=name,
            eid=eid,
            duty=canon_duty(sheet.cell(row, duty_col).value),
            duty_raw=clean_text(sheet.cell(row, duty_col).value),
            group="",
            hire_date=parse_date(sheet.cell(row, hire_col).value),
            row=row,
            workshop=current,
        )
        key = person.key
        if key not in result.frontline:
            result.frontline[key] = person
            result.frontline_order.append(key)
        else:
            result.notes.append(f"「{SHEET_FRONTLINE}」第 {row} 行与前面重复：{name} {eid}")
        if current not in spans:
            spans[current] = [row, row]
            order.append(current)
        else:
            spans[current][1] = row
        result.last_data_row = row

    result.blocks = [(name, spans[name][0], spans[name][1]) for name in order]
    _build_duty_anchors(result)


def _build_duty_anchors(result: BonusFile) -> None:
    """算出每个车间块里每种职务的插入锚点。

    同一职务在一个车间里可能被别的职务隔成几段（示例文件里 11号楼D级车间 的助工
    就被操作工隔成 51 人和 1 人两段），这时取**人数最多**的那一段的末行，
    新人才会跟主群体待在一起，而不是被甩到块尾。
    """
    for workshop, start, end in result.blocks:
        rows = [
            result.frontline[key]
            for key in result.frontline_order
            if start <= result.frontline[key].row <= end
        ]
        runs: list[tuple[str, int, int]] = []  # (职务, 起始行, 末行)
        for person in rows:
            if runs and runs[-1][0] == person.duty and runs[-1][2] == person.row - 1:
                runs[-1] = (person.duty, runs[-1][1], person.row)
            else:
                runs.append((person.duty, person.row, person.row))
        best: dict[str, tuple[int, int]] = {}
        for duty, run_start, run_end in runs:
            size = run_end - run_start + 1
            if duty not in best or size > best[duty][0]:
                best[duty] = (size, run_end)
        for duty, (_, run_end) in best.items():
            result.duty_anchors[(workshop, duty)] = run_end
        seen = []
        for duty, _, _ in runs:
            if duty not in seen:
                seen.append(duty)
        result.duty_order[workshop] = seen
        split = [d for d in seen if sum(1 for r in runs if r[0] == d) > 1]
        if split:
            result.notes.append(
                f"「{workshop}」里 {'、'.join(split)} 被其他职务隔成多段，"
                "新增人员会插到人数最多的那一段末尾"
            )


def _parse_others(sheet) -> SheetLayout:
    """解析「副主任&工艺组长及其他」。

    这张表的车间列**不合并**（每行都写车间名），而且第 130 行之后是混排在
    A~D 列的参照表（产能利用率、基础档位 A/B/C/D，被 L 列的 HLOOKUP 引用），
    所以遇到第一个空姓名行就必须停下，否则会把参照表当成人员。
    """
    from openpyxl.utils import get_column_letter

    header = [cell.value for cell in sheet[1]]
    indexes = {
        "workshop": find_col(header, "车间"),
        "duty": find_col(header, "职务", "岗位"),
        "name": find_col(header, "姓名"),
        "eid": find_col(header, "员工编号", "员工号"),
        "hire": find_col(header, "入职日期", "入职时间"),
    }
    layout = SheetLayout(
        name=sheet.title,
        columns={
            key: get_column_letter(index + 1) for key, index in indexes.items() if index is not None
        },
        first_data_row=2,
        merged_workshop=False,
    )
    order: list[str] = []
    spans: dict[str, list[int]] = {}
    for offset, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        name = norm_name(_get(row, indexes["name"]))
        if not name:
            break
        raw_workshop = _get(row, indexes["workshop"])
        workshop = clean_text(raw_workshop)
        if workshop:
            layout.raw_names.setdefault(workshop, str(raw_workshop).strip())
        person = Person(
            name=name,
            eid=norm_eid(_get(row, indexes["eid"])),
            duty=canon_duty(_get(row, indexes["duty"])),
            duty_raw=clean_text(_get(row, indexes["duty"])),
            group="",
            hire_date=parse_date(_get(row, indexes["hire"])),
            row=offset,
            workshop=workshop,
        )
        layout.people.setdefault(person.key, person)
        layout.last_data_row = offset
        if workshop not in spans:
            spans[workshop] = [offset, offset]
            order.append(workshop)
        else:
            spans[workshop][1] = offset
    layout.blocks = [(name, spans[name][0], spans[name][1]) for name in order]
    _fill_duty_anchors(layout)
    return layout


def _fill_duty_anchors(layout: SheetLayout) -> None:
    by_row = {person.row: person for person in layout.people.values()}
    for workshop, start, end in layout.blocks:
        runs: list[tuple[str, int, int]] = []
        for row in range(start, end + 1):
            person = by_row.get(row)
            if person is None:
                continue
            if runs and runs[-1][0] == person.duty and runs[-1][2] == row - 1:
                runs[-1] = (person.duty, runs[-1][1], row)
            else:
                runs.append((person.duty, row, row))
        best: dict[str, tuple[int, int]] = {}
        for duty, run_start, run_end in runs:
            size = run_end - run_start + 1
            if duty not in best or size > best[duty][0]:
                best[duty] = (size, run_end)
        for duty, (_, run_end) in best.items():
            layout.duty_anchors[(workshop, duty)] = run_end
        seen: list[str] = []
        for duty, _, _ in runs:
            if duty not in seen:
                seen.append(duty)
        layout.duty_order[workshop] = seen


# --------------------------------------------------------------------------- #
# 车间映射
# --------------------------------------------------------------------------- #


@dataclass
class WorkshopGuess:
    group: str
    workshop: str
    source: str
    support: int = 0
    share: float = 0.0
    headcount: int = 0

    @property
    def confidence(self) -> str:
        if self.source == "经验":
            if self.support >= 5 and self.share >= 0.7:
                return "高"
            if self.support >= 2:
                return "中"
            return "低"
        if self.source == "规则":
            return "中"
        return "低"


_PAREN = re.compile(r"[（(]([^（()）]*)[)）]")
_TRAINEE = re.compile(r"[（(]([^（()）]*?)委培[)）]")


def build_workshop_mapping(roster: RosterFile, bonus: BonusFile) -> dict[str, WorkshopGuess]:
    """先用两表已匹配人员反推经验映射，覆盖不到的再用规则兜底。"""
    pairs: dict[str, Counter] = defaultdict(Counter)
    for key, person in bonus.frontline.items():
        source = roster.production.get(key) or roster.production_all.get(key)
        if source is None or not source.group:
            continue
        pairs[source.group][person.workshop] += 1

    headcount = Counter(person.group for person in roster.production.values())
    workshops = set(bonus.workshops)
    mapping: dict[str, WorkshopGuess] = {}
    for group, counter in pairs.items():
        workshop, support = counter.most_common(1)[0]
        total = sum(counter.values())
        mapping[group] = WorkshopGuess(
            group=group,
            workshop=workshop,
            source="经验",
            support=support,
            share=support / total if total else 0.0,
            headcount=headcount.get(group, 0),
        )

    for group in headcount:
        if group in mapping:
            continue
        guessed = _rule_workshop(group, workshops, mapping)
        mapping[group] = WorkshopGuess(
            group=group,
            workshop=guessed or "",
            source="规则" if guessed else "未知",
            headcount=headcount.get(group, 0),
        )
    return dict(sorted(mapping.items(), key=lambda item: -item[1].headcount))


def _rule_workshop(group: str, workshops: set[str], mapping: dict[str, WorkshopGuess]) -> str:
    if not group:
        return ""
    if group in workshops:
        return group
    # X（Y委培）：实际在 Y 干活
    trainee = _TRAINEE.search(group)
    if trainee:
        target = clean_text(trainee.group(1))
        if target in workshops:
            return target
        guess = mapping.get(target)
        if guess and guess.workshop:
            return guess.workshop
    # 去掉括号后缀再试
    base = clean_text(_PAREN.sub("", group))
    if base and base != group:
        if base in workshops:
            return base
        guess = mapping.get(base)
        if guess and guess.workshop:
            return guess.workshop
    # 区域 / 车间 用词差异
    normalized = base.replace("区域", "车间") if base else ""
    if normalized in workshops:
        return normalized
    for workshop in workshops:
        if base and (base in workshop or workshop in base):
            return workshop
    return ""


# --------------------------------------------------------------------------- #
# 对账
# --------------------------------------------------------------------------- #


ACTION_ADD = "add"
ACTION_REMOVE = "remove"
ACTION_UPDATE = "update"
ACTION_MOVE = "move"  # 从「一线人员」移到「副主任&工艺组长及其他」

INTERN_SENIOR = "入职超过3个月"
INTERN_JUNIOR = "入职不到3个月"
INTERN_UNKNOWN = "入职日期异常"
INTERN_MONTHS = 3

# 「待定·核算有清单无」如果能在清单里找到同一个人，改的就是信息而不是删人；
# 这几个字段会按清单值覆盖核算表里的原行。
UPDATE_FIELDS = ("姓名", "员工编号", "职务", "入职日期")


@dataclass
class DiffItem:
    key: tuple[str, str]
    name: str
    eid: str
    category: str
    duty: str
    duty_raw: str
    group: str = ""
    workshop: str = ""
    hire_date: _dt.date | None = None
    leave_date: _dt.date | None = None
    leave_raw: str = ""
    departure_remark: str = ""
    reason: str = ""
    flags: list[str] = field(default_factory=list)
    frontline_row: int = 0
    roster_row: int = 0
    is_intern: bool = False
    intern_source: str = ""
    intern_class: str = ""
    action: str = ACTION_ADD
    updates: dict[str, tuple[str, str]] = field(default_factory=dict)
    new_values: dict[str, object] = field(default_factory=dict)
    target_sheet: str = ""
    target_workshop: str = ""
    target_workshop_source: str = ""

    @property
    def label(self) -> str:
        return key_label(self.key)

    @property
    def update_text(self) -> str:
        return "；".join(f"{name} {old or '空'}→{new or '空'}" for name, (old, new) in self.updates.items())

    @property
    def action_text(self) -> str:
        if self.action == ACTION_MOVE:
            return f"移到「{self.target_sheet}」·{self.target_workshop or '待指定车间'}"
        if self.action == ACTION_UPDATE:
            return "保留在「一线人员」并按清单更新"
        if self.action == ACTION_REMOVE:
            return "从「一线人员」删除"
        return "新增到「一线人员」"

    @property
    def intern_text(self) -> str:
        if not self.is_intern:
            return ""
        return f"{self.intern_class}（据{self.intern_source}）" if self.intern_class else "是"

    def as_dict(self) -> dict:
        return {
            "分类": self.category,
            "姓名": self.name,
            "员工编号": self.eid,
            "职务": self.duty,
            "车间": self.workshop,
            "目前分组": self.group,
            "入职时间": fmt_date(self.hire_date),
            "离职时间": fmt_date(self.leave_date) or self.leave_raw,
            "离职/调出备注": self.departure_remark,
            "清单数据差异": self.update_text,
            "动作": self.action_text,
            "实习生": self.intern_text,
            "实习生分类": self.intern_class,
            "判定依据": self.reason,
            "提示": "；".join(self.flags),
            "_key": self.label,
        }


@dataclass
class Reconciliation:
    items: list[DiffItem]
    ref_date: _dt.date | None
    new_hire_since: _dt.date | None
    mapping: dict[str, WorkshopGuess]
    intern_asof: _dt.date | None = None
    intern_months: int = INTERN_MONTHS
    matched: int = 0
    only_roster: int = 0
    only_bonus: int = 0
    excluded_in_others: int = 0
    paired_renames: int = 0
    notes: list[str] = field(default_factory=list)
    unmapped_groups: list[tuple[str, int]] = field(default_factory=list)

    @property
    def counts(self) -> dict[str, int]:
        counter = Counter(item.category for item in self.items)
        return {category: counter.get(category, 0) for category in CATEGORIES}

    def by_category(self, category: str) -> list[DiffItem]:
        return [item for item in self.items if item.category == category]

    def by_action(self, action: str) -> list[DiffItem]:
        return [item for item in self.items if item.action == action]

    @property
    def intern_counts(self) -> dict[str, int]:
        counter = Counter(item.intern_class for item in self.items if item.is_intern)
        return {
            key: counter.get(key, 0)
            for key in (INTERN_SENIOR, INTERN_JUNIOR, INTERN_UNKNOWN)
            if counter.get(key, 0)
        }

    def category_of(self, name: str, eid: str) -> str | None:
        target = (norm_name(name), norm_eid(eid))
        for item in self.items:
            if item.key == target:
                return item.category
        return None


def _match_in_roster(key, roster: RosterFile, by_eid, by_name) -> Person | None:
    """在整张「生产部」里找同一个人：先精确，再编号唯一，最后姓名唯一。"""
    exact = roster.production_all.get(key)
    if exact is not None:
        return exact
    candidates = by_eid.get(key[1], [])
    if len(candidates) == 1:
        return candidates[0]
    candidates = by_name.get(key[0], [])
    if len(candidates) == 1:
        return candidates[0]
    return None


def _attach_update(
    item: DiffItem, current: Person, source: Person, others_workshops: dict[str, str]
) -> None:
    """把"按清单数据更新"所需的新值与差异说明挂到 item 上。

    更新后如果职务已经不属于一线的四种职务，这个人就该挪到
    「副主任&工艺组长及其他」子表，动作从 update 变成 move。
    """
    pairs = (
        ("姓名", current.name, source.name),
        ("员工编号", current.eid, source.eid),
        ("职务", current.duty, source.duty),
        ("入职日期", fmt_date(current.hire_date), fmt_date(source.hire_date)),
    )
    item.new_values = {
        "姓名": source.name,
        "员工编号": source.eid,
        "职务": source.duty,
        "入职日期": source.hire_date,
    }
    item.updates = {name: (old, new) for name, old, new in pairs if old != new}
    item.is_intern = source.is_intern
    item.intern_source = source.intern_source
    item.roster_row = source.row
    item.reason += (
        f"；「{SHEET_PRODUCTION}」第 {source.row} 行有同一人"
        f"（{source.name} {source.eid} 职位「{source.title or source.duty_raw}」）"
    )
    if is_target_duty(source.duty_raw):
        item.action = ACTION_UPDATE
        item.reason += "，职务仍属一线四种职务，按清单更新后保留"
        return
    item.action = ACTION_MOVE
    item.target_sheet = SHEET_OTHERS
    mapped = others_workshops.get(source.group)
    item.target_workshop = mapped or current.workshop
    item.target_workshop_source = "经验映射" if mapped else "沿用一线车间"
    item.reason += f"，职务「{source.duty}」不属一线四种职务，移到「{SHEET_OTHERS}」"


def new_hire_evidence(
    person: Person,
    roster: RosterFile,
    new_hire_since: _dt.date | None,
    anchor: _dt.date | None,
) -> tuple[bool, list[str], list[str]]:
    """判断"清单有、目标表无"的人算不算新入职，返回 (是否新入职, 依据, 提示)。

    两个功能（一线人员 / 副主任&工艺组长及其他）共用同一套口径。
    """
    reasons: list[str] = []
    flags: list[str] = []
    change = roster.changes.get(person.key)
    is_new = (
        person.hire_date is not None
        and new_hire_since is not None
        and person.hire_date >= new_hire_since
        and (anchor is None or person.hire_date < anchor)
    )
    if is_new:
        reasons.append(f"入职 {person.hire_date} 落在 {new_hire_since} ~ {anchor} 之间")
    if change is not None and change["leave_blank"]:
        is_new = True
        reasons.append(f"「{SHEET_CHANGES}」第 {change['row']} 行（{change['kind']}）且离职时间为空")
    elif change is not None:
        reasons.append(f"「{SHEET_CHANGES}」第 {change['row']} 行已有离职时间 {change['leave_raw']}")
    if person.hire_date and anchor and person.hire_date >= anchor:
        flags.append(f"入职时间 {person.hire_date} 不早于参照日期 {anchor}")
    if not reasons:
        reasons.append(
            f"入职 {fmt_date(person.hire_date) or '未知'} 不在 {new_hire_since} ~ {anchor} 之间，"
            f"且「{SHEET_CHANGES}」无在职记录"
        )
    return is_new, reasons, flags


def departure_evidence(
    key: tuple[str, str], roster: RosterFile
) -> tuple[bool, list[str], Person | None, dict | None]:
    """判断"目标表有、清单无"的人算不算离职，返回 (是否离职, 依据, 离职表记录, 变动说明记录)。"""
    departure = roster.departures.get(key)
    change = roster.changes.get(key)
    reasons: list[str] = []
    is_left = False
    if departure is not None:
        is_left = True
        reasons.append(
            f"「{SHEET_DEPARTURE}」第 {departure.row} 行"
            + (
                f"，离职时间 {fmt_date(departure.leave_date) or departure.leave_raw}"
                if departure.leave_raw
                else ""
            )
        )
    if change is not None and not change["leave_blank"]:
        is_left = True
        reasons.append(f"「{SHEET_CHANGES}」第 {change['row']} 行有离职时间 {change['leave_raw']}")
    elif change is not None:
        reasons.append(f"「{SHEET_CHANGES}」第 {change['row']} 行（{change['kind']}）离职时间为空")
    if not reasons:
        reasons.append(f"「{SHEET_DEPARTURE}」与「{SHEET_CHANGES}」都查不到离职记录")
    return is_left, reasons, departure, change


def classify_intern(item: DiffItem, asof: _dt.date | None, months: int) -> None:
    _classify_intern(item, asof, months)


def _classify_intern(item: DiffItem, asof: _dt.date | None, months: int) -> None:
    """按判断日期把实习生分成"入职超过 N 个月"和"入职不到 N 个月"。"""
    if not item.is_intern:
        return
    if item.hire_date is None or asof is None or item.hire_date >= asof:
        item.intern_class = INTERN_UNKNOWN
        item.flags.append(
            f"实习生（据{item.intern_source}），入职 {fmt_date(item.hire_date) or '未知'} "
            f"不早于判断日期 {asof}，无法计算入职时长"
        )
        return
    cutoff = months_before(asof, months)
    item.intern_class = INTERN_SENIOR if item.hire_date <= cutoff else INTERN_JUNIOR
    item.flags.append(
        f"实习生（据{item.intern_source}），{item.intern_class}"
        f"（按 {asof} 算，{months} 个月分界为 {cutoff}）"
    )


def build_others_workshop_map(roster: RosterFile, bonus: BonusFile) -> dict[str, str]:
    """「目前分组」→「副主任&工艺组长及其他」的车间名。

    两张子表的车间叫法不一样（一线人员写「11号楼D级车间」，这张表写「11号楼车间D级区域」），
    所以同样用两表已匹配的人反推，而不是猜。
    """
    layout = bonus.others_layout
    if layout is None:
        return {}
    votes: dict[str, Counter] = defaultdict(Counter)
    for key, person in layout.people.items():
        source = roster.production_all.get(key)
        if source is not None and source.group and person.workshop:
            votes[source.group][person.workshop] += 1
    return {group: counter.most_common(1)[0][0] for group, counter in votes.items()}


def reconcile(
    roster: RosterFile,
    bonus: BonusFile,
    *,
    ref_date: _dt.date | None = None,
    new_hire_since: _dt.date | None = None,
    intern_asof: _dt.date | None = None,
    intern_months: int = INTERN_MONTHS,
    exclude_in_others: bool = True,
    mapping: dict[str, WorkshopGuess] | None = None,
) -> Reconciliation:
    """对账。

    ``new_hire_since``：新入职判定窗口的起点；入职时间需落在
    ``[new_hire_since, ref_date]`` 之间才算新入职。
    ``intern_asof``：实习生"入职是否满 N 个月"的判断日期，默认取参照日期。
    """
    anchor = ref_date or roster.ref_date
    if new_hire_since is None:
        new_hire_since = months_before(anchor, 1) if anchor else None
    if intern_asof is None:
        intern_asof = anchor
    mapping = mapping or build_workshop_mapping(roster, bonus)
    others_workshops = build_others_workshop_map(roster, bonus)

    roster_keys = set(roster.production)
    bonus_keys = set(bonus.frontline)
    matched = roster_keys & bonus_keys

    eid_to_bonus = defaultdict(list)
    name_to_bonus = defaultdict(list)
    for key in bonus_keys:
        eid_to_bonus[key[1]].append(key)
        name_to_bonus[key[0]].append(key)
    eid_to_roster = defaultdict(list)
    name_to_roster = defaultdict(list)
    for key in roster_keys:
        eid_to_roster[key[1]].append(key)
        name_to_roster[key[0]].append(key)

    # 找"同一个人"要在整张生产部表里找，而不是只在目标职务里找——
    # 职务变动（班长→工艺组长）恰恰是这类差异最常见的成因
    eid_to_roster_all = defaultdict(list)
    name_to_roster_all = defaultdict(list)
    for key, person in roster.production_all.items():
        eid_to_roster_all[key[1]].append(person)
        name_to_roster_all[key[0]].append(person)

    items: list[DiffItem] = []
    excluded = 0

    # --- 清单有、核算无 ------------------------------------------------- #
    for key in sorted(roster_keys - bonus_keys, key=lambda k: (roster.production[k].group, k)):
        person = roster.production[key]
        flags: list[str] = []
        if exclude_in_others and key in bonus.others:
            excluded += 1
            continue
        if key in bonus.others:
            flags.append(f"已在「{SHEET_OTHERS}」子表")
        for other in eid_to_bonus.get(key[1], []):
            if other != key:
                flags.append(f"核算表有同编号不同名：{other[0]}")
        for other in name_to_bonus.get(key[0], []):
            if other != key:
                flags.append(f"核算表有同名不同编号：{other[1]}")

        is_new, reasons, extra_flags = new_hire_evidence(person, roster, new_hire_since, anchor)
        flags.extend(extra_flags)
        guess = mapping.get(person.group)
        item = DiffItem(
            key=key,
            name=person.name,
            eid=person.eid,
            category=CATEGORY_NEW if is_new else CATEGORY_PENDING_ADD,
            duty=person.duty,
            duty_raw=person.duty_raw,
            group=person.group,
            workshop=guess.workshop if guess else "",
            hire_date=person.hire_date,
            departure_remark=person.remark,
            reason="；".join(reasons),
            flags=flags,
            roster_row=person.row,
            is_intern=person.is_intern,
            intern_source=person.intern_source,
            action=ACTION_ADD,
        )
        _classify_intern(item, intern_asof, intern_months)
        items.append(item)

    # --- 核算有、清单无 ------------------------------------------------- #
    for key in sorted(bonus_keys - roster_keys, key=lambda k: bonus.frontline[k].row):
        person = bonus.frontline[key]
        flags = []
        is_left, reasons, departure, change = departure_evidence(key, roster)
        still = roster.production_all.get(key)
        if still is not None:
            flags.append(f"仍在「{SHEET_PRODUCTION}」但职务为「{still.duty_raw}」（非目标职务）")
        if key in roster.equipment:
            flags.append(f"在「{SHEET_EQUIPMENT}」子表")
        for other in eid_to_roster.get(key[1], []):
            if other != key:
                flags.append(f"清单有同编号不同名：{other[0]}")
        for other in name_to_roster.get(key[0], []):
            if other != key:
                flags.append(f"清单有同名不同编号：{other[1]}")

        category = CATEGORY_LEFT if is_left else CATEGORY_PENDING_DEL
        item = DiffItem(
            key=key,
            name=person.name,
            eid=person.eid,
            category=category,
            duty=person.duty,
            duty_raw=person.duty_raw,
            group=departure.group if departure else (still.group if still else ""),
            workshop=person.workshop,
            hire_date=person.hire_date,
            leave_date=departure.leave_date if departure else (change["leave_date"] if change else None),
            leave_raw=departure.leave_raw if departure else (change["leave_raw"] if change else ""),
            departure_remark=departure.remark if departure else (change["remark"] if change else ""),
            reason="；".join(reasons),
            flags=flags,
            frontline_row=person.row,
            action=ACTION_REMOVE,
        )
        if category == CATEGORY_PENDING_DEL:
            # 这类人多半没离职，只是职务变了或姓名录错；能在清单里找到本人就改成"更新信息"
            source = _match_in_roster(key, roster, eid_to_roster_all, name_to_roster_all)
            if source is not None:
                _attach_update(item, person, source, others_workshops)
                _classify_intern(item, intern_asof, intern_months)
        items.append(item)

    # 同一个人可能同时落在两边：清单里叫「曹睿晟」、核算表里叫「曹静旺」（同一个员工编号）。
    # 更新那一侧改完名字就等于这个人已经在表里了，新增那一侧必须撤掉，否则会出现重复行。
    rename_targets = {
        (item.new_values.get("姓名"), item.new_values.get("员工编号")): item
        for item in items
        if item.action == ACTION_UPDATE
    }
    paired = 0
    kept_items = []
    for item in items:
        target = rename_targets.get(item.key) if item.action == ACTION_ADD else None
        if target is not None and target is not item:
            target.flags.append(
                f"与「{SHEET_PRODUCTION}」第 {item.roster_row} 行是同一人，"
                "按清单改名即可，不需要另外新增"
            )
            paired += 1
            continue
        kept_items.append(item)
    items = kept_items

    notes = list(roster.notes) + list(bonus.notes)
    if paired:
        notes.append(
            f"有 {paired} 人在两表里编号相同、姓名不同，已合并为「按清单更新姓名」，不再重复新增"
        )
    pending = Counter(
        item.group
        for item in items
        if item.action == "add" and not (mapping.get(item.group) and mapping[item.group].workshop)
    )
    unmapped = sorted(pending.items(), key=lambda kv: (-kv[1], kv[0]))

    return Reconciliation(
        items=items,
        ref_date=anchor,
        new_hire_since=new_hire_since,
        intern_asof=intern_asof,
        intern_months=intern_months,
        mapping=mapping,
        matched=len(matched),
        only_roster=len(roster_keys - bonus_keys),
        only_bonus=len(bonus_keys - roster_keys),
        excluded_in_others=excluded,
        paired_renames=paired,
        notes=notes,
        unmapped_groups=unmapped,
    )
