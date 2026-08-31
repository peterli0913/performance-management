"""一键生成：同一份核算表同时改「一线人员」和「副主任&工艺组长及其他」。"""

import io
import zipfile

import openpyxl
import pytest

from tj4tools.bonus_export import build_combined_workbook, build_workbook, merge_others_inserts
from tj4tools.roster import ACTION_ADD, ACTION_MOVE, DiffItem
from tj4tools.roster import reconcile
from tj4tools.supervisor import SCOPE_STRICT, reconcile_supervisors

FRONTLINE = "一线人员"
OTHERS = "副主任&工艺组长及其他"


@pytest.fixture(scope="module")
def pair(roster_with_interns, bonus):
    frontline = reconcile(roster_with_interns, bonus)
    from tj4tools.roster import placeable_keys

    placeable = placeable_keys(frontline)
    supervisor = reconcile_supervisors(
        roster_with_interns, bonus, scope=SCOPE_STRICT, placeable_keys=placeable
    )
    return frontline, supervisor


def _item(name, eid, action, duty="", workshop=""):
    return DiffItem(
        key=(name, eid),
        name=name,
        eid=eid,
        category="测试",
        duty=duty,
        duty_raw=duty,
        workshop=workshop,
        action=action,
        target_workshop=workshop,
        new_values={"职务": duty} if action == ACTION_MOVE else {},
    )


def test_merge_others_inserts_prefers_supervisor_item_and_does_not_mutate():
    move_only = _item("甲", "1", ACTION_MOVE, "清洗工", "清洗组")
    extra_only = _item("乙", "2", ACTION_ADD, "助理工程师", "安全组")
    move_overlap = _item("丙", "3", ACTION_MOVE, "车间副主任", "12号楼")
    extra_overlap = _item("丙", "3", ACTION_ADD, "副主任", "")
    merged = merge_others_inserts([move_only, move_overlap], [extra_only, extra_overlap])
    assert merged == [move_only, extra_overlap, extra_only]
    assert extra_overlap.target_workshop == ""
    assert extra_overlap.workshop == ""


def _payload(pair):
    frontline, supervisor = pair
    return {
        "frontline_adds": [i for i in frontline.items if i.action == "add" and i.workshop],
        "frontline_removes": [i for i in frontline.items if i.action == "remove"],
        "frontline_updates": [i for i in frontline.items if i.action == "update"],
        "frontline_moves": [i for i in frontline.items if i.action == "move"],
        "supervisor_adds": [i for i in supervisor.items if i.action == "add"],
        "supervisor_removes": [i for i in supervisor.items if i.action == "remove"],
    }


@pytest.fixture(scope="module")
def combined(bonus_bytes, bonus, pair):
    data, summary = build_combined_workbook(bonus_bytes, bonus, **_payload(pair), mode="apply")
    return data, summary, _payload(pair)


def test_combined_workbook_is_not_the_frontline_only_export(bonus_bytes, bonus, pair):
    """还没有一键导出时，这个符号就必须不存在——用来锁住 API。"""
    payload = _payload(pair)
    data, summary = build_combined_workbook(bonus_bytes, bonus, **payload, mode="apply")
    only_frontline, _ = build_workbook(
        bonus_bytes,
        bonus,
        payload["frontline_adds"],
        payload["frontline_removes"],
        payload["frontline_updates"],
        payload["frontline_moves"],
        mode="apply",
    )
    assert data != only_frontline
    assert summary.other_added > 0
    assert summary.other_removed > 0


def test_combined_apply_changes_both_sheets(combined, bonus):
    data, summary, payload = combined
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    frontline = workbook[FRONTLINE]
    others = workbook[OTHERS]

    move_keys = {item.key for item in payload["frontline_moves"]}
    unique_supervisor = [
        item
        for item in payload["supervisor_adds"]
        if item.key not in move_keys and (item.target_workshop or item.workshop)
    ]
    # 一线 228 增 / 41 删 / 15 移；转移组等改走副主任表后，去重新增 16、删除 11
    assert summary.added == len(payload["frontline_adds"]) == 228
    assert summary.removed == len(payload["frontline_removes"]) == 41
    assert summary.moved == 15
    assert summary.updated == 1
    assert summary.other_removed == 11
    assert summary.other_added == len(unique_supervisor) == 16
    assert frontline.max_row == 676 + summary.added - summary.removed - summary.moved
    assert others.max_row == 206 + summary.moved + summary.other_added - summary.other_removed


def test_overlapping_people_land_once_with_supervisor_duty(combined, pair):
    """6 人既是一线移出、又是副主任表待新增：只能插一行，职务用副主任表写法。"""
    data, _, payload = combined
    frontline, supervisor = pair
    move_keys = {item.key for item in payload["frontline_moves"]}
    overlap = [item for item in supervisor.items if item.action == "add" and item.key in move_keys]
    assert len(overlap) == 6

    sheet = openpyxl.load_workbook(io.BytesIO(data))[OTHERS]
    landed: dict[str, list[tuple[int, str]]] = {}
    for row in range(2, sheet.max_row + 1):
        name = str(sheet.cell(row, 3).value or "")
        if not name:
            break
        landed.setdefault(name, []).append((row, str(sheet.cell(row, 2).value or "").strip()))

    for item in overlap:
        hits = landed[item.name]
        assert len(hits) == 1, item.name
        assert hits[0][1] == item.duty, (item.name, hits[0][1], item.duty)

    # 车间空的副主任新增，要沿用一线移出时的目标车间，不能被去重后丢掉
    empty = [item for item in overlap if not (item.target_workshop or item.workshop)]
    assert {item.name for item in empty} == {"吴宗璇", "王腾蛟"}
    move_by_key = {item.key: item for item in payload["frontline_moves"]}
    for item in empty:
        row, _ = landed[item.name][0]
        assert str(sheet.cell(row, 1).value or "").strip() == move_by_key[item.key].target_workshop


def test_combined_keeps_other_sheets_and_parts(bonus_bytes, combined):
    data, _, _ = combined
    before = set(zipfile.ZipFile(io.BytesIO(bonus_bytes)).namelist())
    after = set(zipfile.ZipFile(io.BytesIO(data)).namelist())
    assert before - after == {"xl/calcChain.xml"}
    assert "xl/externalLinks/externalLink1.xml" in after

    src = openpyxl.load_workbook(io.BytesIO(bonus_bytes), data_only=False)
    dst = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    assert src.sheetnames == dst.sheetnames
    for name in src.sheetnames:
        if name in (FRONTLINE, OTHERS):
            continue
        left, right = src[name], dst[name]
        assert left.max_row == right.max_row, name
        for row in left.iter_rows():
            for cell in row:
                assert right[cell.coordinate].value == cell.value, (name, cell.coordinate)


def test_combined_lookup_reference_follows_net_delta(bonus_bytes, combined):
    data, _, _ = combined
    sheet = openpyxl.load_workbook(io.BytesIO(data), data_only=False)[OTHERS]
    # 15 移入 + 16 副主任新增 - 11 删除 = +20，不能用 summary 反推，否则摘要错了公式也会过
    assert sheet["L2"].value == "=HLOOKUP(M2,$A$163:$H$164,2)"


def test_combined_others_workshop_blocks_stay_contiguous(combined):
    sheet = openpyxl.load_workbook(io.BytesIO(combined[0]))[OTHERS]
    sequence = []
    for row in range(2, sheet.max_row + 1):
        if not sheet.cell(row, 3).value:
            break
        sequence.append(str(sheet.cell(row, 1).value or "").strip())
    runs = [v for index, v in enumerate(sequence) if index == 0 or sequence[index - 1] != v]
    assert len(runs) == len(set(runs)), runs


def test_combined_summary_mentions_both_sheets(combined):
    text = combined[1].text()
    assert "直接插入 228 人（红字）" in text
    assert "直接删除 41 人" in text
    assert "副主任表直接插入 16 人（红字）" in text
    assert "副主任表直接删除 11 人" in text
    assert "移到「副主任&工艺组长及其他」15 人" in text
