"""功能二：「副主任&工艺组长及其他」子表的取人口径与对账。"""

import datetime as dt
import io
import zipfile

import openpyxl
import pytest

from tj4tools.roster import (
    CATEGORY_LEFT,
    CATEGORY_NEW,
    CATEGORY_PENDING_ADD,
    CATEGORY_PENDING_DEL,
    reconcile,
)
from tj4tools.supervisor import (
    FRONTLINE_TITLES,
    MANAGEMENT_TITLES,
    SCOPE_LITERAL,
    SCOPE_STRICT,
    build_duty_map,
    build_target,
    reconcile_supervisors,
    target_summary,
)
from tj4tools.supervisor_export import build_supervisor_workbook, split_by_action

SHEET = "副主任&工艺组长及其他"
FRONTLINE = "一线人员"


@pytest.fixture(scope="module")
def placeable(roster_with_interns, bonus):
    """功能一能放进一线车间的人——严格口径要把他们排除掉。"""
    analysis = reconcile(roster_with_interns, bonus)
    return {item.key for item in analysis.items if item.action == "add" and item.workshop}


@pytest.fixture(scope="module")
def strict(roster_with_interns, bonus, placeable):
    return reconcile_supervisors(
        roster_with_interns, bonus, scope=SCOPE_STRICT, placeable_keys=placeable
    )


def test_duty_map_is_derived_from_matched_people(roster_with_interns, bonus):
    mapping = build_duty_map(roster_with_interns, bonus)
    # 清单写「车间副主任」，这张表写「副主任」
    assert mapping["车间副主任"] == "副主任"
    assert mapping["助理工程师"] == "助理工程师"
    assert mapping["工艺组长"] == "工艺组长"
    assert mapping["工程师"] == "工程师"
    assert mapping["经理"] == "经理"
    # 兜底项：清单里有、表里还没出现过的写法
    assert mapping["工艺副主管"] == "工艺组长"
    assert mapping["安全员"] == "助工"


def test_two_scopes_differ_by_an_order_of_magnitude(roster_with_interns, bonus, placeable):
    totals = target_summary(roster_with_interns, bonus, placeable)
    assert totals[SCOPE_STRICT] == 150
    assert totals[SCOPE_LITERAL] == 387
    assert totals[SCOPE_LITERAL] > totals[SCOPE_STRICT] * 2


def test_target_is_management_plus_leftover_frontline(roster_with_interns, bonus, placeable):
    target, groups, _ = build_target(
        roster_with_interns, bonus, scope=SCOPE_STRICT, placeable_keys=placeable
    )
    management = [k for k, g in groups.items() if g == "管理类职务"]
    leftover = [k for k, g in groups.items() if g != "管理类职务"]
    assert len(management) == 68
    assert len(leftover) == 82
    assert all(target[k].title in MANAGEMENT_TITLES for k in management)
    assert all(target[k].title in FRONTLINE_TITLES for k in leftover)
    # 组二的人一定不在一线人员子表里
    assert all(k not in bonus.frontline for k in leftover)
    # 组一不看是否在一线人员表
    assert any(k in bonus.frontline for k in management)


def test_strict_scope_excludes_people_feature_one_can_place(roster_with_interns, bonus, placeable):
    strict_target, _, _ = build_target(
        roster_with_interns, bonus, scope=SCOPE_STRICT, placeable_keys=placeable
    )
    literal_target, _, _ = build_target(
        roster_with_interns, bonus, scope=SCOPE_LITERAL, placeable_keys=placeable
    )
    extra = set(literal_target) - set(strict_target)
    assert extra
    # 字面口径多出来的人正是功能一要放进一线人员的那批，两个功能会抢人
    assert extra <= placeable


def test_strict_reconciliation_shape(strict, bonus):
    assert strict.matched == 117
    assert strict.only_roster == 33
    assert strict.only_bonus == 11
    assert sum(strict.counts.values()) == strict.only_roster + strict.only_bonus
    assert strict.counts[CATEGORY_NEW] == 20
    assert strict.counts[CATEGORY_PENDING_ADD] == 13
    assert strict.counts[CATEGORY_LEFT] == 6
    assert strict.counts[CATEGORY_PENDING_DEL] == 5
    assert any("取人口径" in note for note in strict.notes)


def test_adds_use_this_sheet_duty_and_workshop_naming(strict):
    adds = [item for item in strict.items if item.action == "add"]
    assert adds
    promoted = next(i for i in adds if i.duty_raw == "车间副主任")
    assert promoted.duty == "副主任", "职务要转成本表的写法"
    mapped = [i for i in adds if i.workshop]
    assert mapped
    # 本表车间名和一线人员不同
    assert any(i.workshop == "11号楼车间D级区域" for i in mapped)
    assert all(i.target_sheet == "副主任&工艺组长及其他" for i in adds)


def test_removes_explain_why_they_are_out_of_scope(strict):
    removes = [item for item in strict.items if item.action == "remove"]
    assert len(removes) == 11
    left = [i for i in removes if i.category == CATEGORY_LEFT]
    pending = [i for i in removes if i.category == CATEGORY_PENDING_DEL]
    assert len(left) == 6 and len(pending) == 5
    for item in left:
        assert "离职人员&调出人员" in item.reason or "有离职时间" in item.reason
    # 口径外职务：工艺副主管 3 人 + 安全员 2 人
    titles = sorted(
        flag.split("「")[1].split("」")[0]
        for item in pending
        for flag in item.flags
        if "清单职位为" in flag
    )
    assert titles == ["安全员", "安全员", "工艺副主管", "工艺副主管", "工艺副主管"]


def test_unmapped_groups_are_reported(strict):
    """只给出数据，提示语交给界面按"当前生效的映射"实时算，避免人工指定完文案还挂着。"""
    assert strict.unmapped_groups
    assert sum(count for _, count in strict.unmapped_groups) == 14
    assert not any("没有对应车间" in note for note in strict.notes)
    unmapped = [i for i in strict.items if i.action == "add" and not i.workshop]
    assert len(unmapped) == 14
    assert all(any("没有对应车间" in flag for flag in i.flags) for i in unmapped)


def test_literal_scope_pulls_in_the_frontline_backlog(roster_with_interns, bonus, placeable):
    literal = reconcile_supervisors(
        roster_with_interns, bonus, scope=SCOPE_LITERAL, placeable_keys=placeable
    )
    assert literal.only_roster == 270
    assert literal.only_bonus == 11
    assert literal.counts[CATEGORY_NEW] > 200


def test_intern_classification_carries_over(roster_with_interns, bonus, placeable):
    strict = reconcile_supervisors(
        roster_with_interns,
        bonus,
        scope=SCOPE_STRICT,
        placeable_keys=placeable,
        intern_asof=dt.date(2026, 12, 31),
    )
    interns = [i for i in strict.items if i.is_intern]
    assert interns
    assert all(i.intern_class for i in interns)


# --------------------------------------------------------------------------- #
# 导出
# --------------------------------------------------------------------------- #


@pytest.fixture(scope="module")
def generated(bonus_bytes, bonus, strict):
    items = [i for i in strict.items if i.action != "add" or i.workshop]
    adds, removes = split_by_action(items)
    mark, mark_summary = build_supervisor_workbook(bonus_bytes, bonus, adds, removes, mode="mark")
    applied, applied_summary = build_supervisor_workbook(
        bonus_bytes, bonus, adds, removes, mode="apply"
    )
    return {
        "adds": adds,
        "removes": removes,
        "mark": (mark, mark_summary),
        "apply": (applied, applied_summary),
    }


def test_export_only_touches_this_sheet(bonus_bytes, generated):
    data, _ = generated["apply"]
    before = openpyxl.load_workbook(io.BytesIO(bonus_bytes), data_only=False)
    after = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    assert before.sheetnames == after.sheetnames
    for name in before.sheetnames:
        if name == SHEET:
            continue
        source, target = before[name], after[name]
        assert source.max_row == target.max_row, name
        for row in source.iter_rows():
            for cell in row:
                assert target[cell.coordinate].value == cell.value, (name, cell.coordinate)


def test_export_keeps_every_part(bonus_bytes, generated):
    data, _ = generated["apply"]
    before = set(zipfile.ZipFile(io.BytesIO(bonus_bytes)).namelist())
    after = set(zipfile.ZipFile(io.BytesIO(data)).namelist())
    assert before - after == {"xl/calcChain.xml"}
    assert "xl/externalLinks/externalLink1.xml" in after
    assert "xl/comments1.xml" in after  # 这张表上的批注


def test_apply_mode_row_count(generated):
    data, summary = generated["apply"]
    sheet = openpyxl.load_workbook(io.BytesIO(data))[SHEET]
    assert summary.added == len(generated["adds"])
    assert summary.removed == len(generated["removes"])
    assert sheet.max_row == 206 + summary.added - summary.removed


def test_workshop_text_is_written_verbatim(bonus_bytes, bonus, roster_with_interns, placeable):
    """「生产技术转移组（多肽）」是全角括号，写回必须一字不差，否则车间块会被拆成两段。"""
    layout = bonus.others_layout
    assert layout.raw_names["生产技术转移组(多肽)"] == "生产技术转移组（多肽）"
    strict = reconcile_supervisors(
        roster_with_interns, bonus, scope=SCOPE_STRICT, placeable_keys=placeable
    )
    adds = [i for i in strict.items if i.action == "add" and i.workshop == "生产技术转移组(多肽)"]
    assert adds, "样本里应当有要加进这个车间的人"
    data, _ = build_supervisor_workbook(bonus_bytes, bonus, adds, [], mode="apply")
    sheet = openpyxl.load_workbook(io.BytesIO(data))[SHEET]
    written = {
        str(sheet.cell(row, 1).value or "")
        for row in range(2, sheet.max_row + 1)
        if str(sheet.cell(row, 3).value or "") in {i.name for i in adds}
    }
    assert written == {"生产技术转移组（多肽）"}


def test_workshop_blocks_stay_contiguous(generated):
    for mode in ("mark", "apply"):
        data, _ = generated[mode]
        sheet = openpyxl.load_workbook(io.BytesIO(data))[SHEET]
        sequence = []
        for row in range(2, sheet.max_row + 1):
            if not sheet.cell(row, 3).value:
                break
            sequence.append(str(sheet.cell(row, 1).value or "").strip())
        runs = [v for index, v in enumerate(sequence) if index == 0 or sequence[index - 1] != v]
        assert len(runs) == len(set(runs)), (mode, runs)


def test_inserted_rows_carry_row_formulas(generated):
    """公式要指向自己这一行。注意这张表不同车间块的 F 列公式并不完全一样
    （生产技术转移组那几块带 `/30*20`），所以只断言行号而不锁死整个公式。"""
    import re

    data, _ = generated["apply"]
    sheet = openpyxl.load_workbook(io.BytesIO(data), data_only=False)[SHEET]
    names = {i.name for i in generated["adds"]}
    pattern = re.compile(r"(?<![A-Za-z0-9_$.!])\$?[A-Z]{1,3}\$?(\d{1,7})(?![\d(])")
    checked = 0
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row, 3).value or "") not in names:
            continue
        formula = sheet.cell(row, 6).value
        assert formula.startswith(f"=(L{row}*K{row}"), (row, formula)
        assert {int(n) for n in pattern.findall(formula)} == {row}, (row, formula)
        assert f"D{row}" in sheet.cell(row, 18).value
        assert "[1]生产部" in sheet.cell(row, 18).value
        assert sheet.cell(row, 5).is_date
        checked += 1
    assert checked == len(names)


def test_lookup_table_reference_follows(bonus_bytes, generated):
    source = openpyxl.load_workbook(io.BytesIO(bonus_bytes), data_only=False)[SHEET]
    assert source["L2"].value == "=HLOOKUP(M2,$A$143:$H$144,2)"
    data, summary = generated["apply"]
    sheet = openpyxl.load_workbook(io.BytesIO(data), data_only=False)[SHEET]
    delta = summary.added - summary.removed
    assert sheet["L2"].value == f"=HLOOKUP(M2,$A${143 + delta}:$H${144 + delta},2)"


def _rows_by_name(sheet):
    out = {}
    for row in range(2, sheet.max_row + 1):
        name = str(sheet.cell(row, 3).value or "")
        if name:
            out.setdefault(name, row)
    return out


def test_mark_mode_colors(generated):
    data, _ = generated["mark"]
    sheet = openpyxl.load_workbook(io.BytesIO(data))[SHEET]
    assert sheet.max_row == 206 + len(generated["adds"]), "标记版不删行"
    # 插入会把原有行往下推，所以要按姓名找新行号，不能用原始行号
    rows = _rows_by_name(sheet)
    for item in generated["removes"]:
        row = rows[item.name]
        assert sheet.cell(row, 3).fill.fgColor.rgb == "FFFF0000", item.name
        assert sheet.cell(row, 4).fill.fgColor.rgb == "FFFF0000", item.name
    for item in generated["adds"]:
        row = rows[item.name]
        # 实习生走黄底，其余新增走绿底
        expected = "FFFFFF00" if item.is_intern else "FF92D050"
        assert sheet.cell(row, 3).fill.fgColor.rgb == expected, item.name


def test_apply_mode_uses_red_font(generated):
    """非实习生的新增行只改字体颜色，字体名和字号沿用同一段的原有行
    （这张表各块字体并不统一）；实习生不叠红字，走整片黄底。"""
    data, _ = generated["apply"]
    sheet = openpyxl.load_workbook(io.BytesIO(data))[SHEET]
    by_name = {i.name: i for i in generated["adds"]}
    checked = interns = 0
    for row in range(2, sheet.max_row + 1):
        item = by_name.get(str(sheet.cell(row, 3).value or ""))
        if item is None:
            continue
        cell = sheet.cell(row, 3)
        above = sheet.cell(row - 1, 3)
        if item.is_intern:
            assert cell.fill.fgColor.rgb == "FFFFFF00", item.name
            assert cell.font.color is None or cell.font.color.rgb != "FFFF0000", item.name
            interns += 1
        else:
            assert cell.font.color is not None and cell.font.color.rgb == "FFFF0000"
            assert cell.font.name == above.font.name, row
            assert cell.font.sz == above.font.sz, row
            assert cell.fill.fgColor.rgb != "FF92D050"
        checked += 1
    assert checked == len(by_name)
    assert interns >= 1, "样本里应当有实习生走黄底这条分支"
