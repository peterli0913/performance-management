from __future__ import annotations

import io
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from tj4tools.bonus_export import build_workbook, split_by_action
from tj4tools.ingest import ingest_files
from tj4tools.roster import (
    ACTION_ADD,
    ACTION_MOVE,
    ACTION_REMOVE,
    ACTION_UPDATE,
    SHEET_FRONTLINE,
    SHEET_OTHERS,
    Person,
    parse_bonus,
)
from tj4tools.search import (
    CATEGORY_MANUAL,
    ManualPlacement,
    search_tables,
    split_placements,
)


SAMPLES = Path(__file__).resolve().parent.parent
ROSTER = SAMPLES / "TJ4生产部&生产设备部人员清单，07-31-2026.xlsx"
BONUS = SAMPLES / "2026年07月份安全质量奖核算数据.xlsx"


def _tables():
    items = [(path.name, path.read_bytes()) for path in (ROSTER, BONUS)]
    return ingest_files(items).tables


def _bonus():
    return parse_bonus(BONUS.read_bytes(), BONUS.name)


def test_empty_query_returns_nothing():
    assert search_tables(_tables(), "") == []
    assert search_tables(_tables(), "   ") == []


def test_search_by_name_shows_file_sheet_and_full_row():
    hits = search_tables(_tables(), "张文艺")
    assert hits, "黄金样本里应能搜到张文艺"
    hit = next(h for h in hits if "人员清单" in h.file_name)
    assert hit.sheet_name
    assert hit.excel_row >= 2
    assert hit.cells, "必须带回整行原文"
    assert any("张文艺" in v for v in hit.cells.values())
    assert hit.name == "张文艺"
    assert hit.eid


def test_search_by_employee_id():
    hits = search_tables(_tables(), "ALS0702")
    assert any(h.eid == "ALS0702" or "ALS0702" in h.cells.values() for h in hits)


def test_search_by_keyword_matches_any_cell():
    hits = search_tables(_tables(), "委培")
    assert hits
    assert all(h.matched for h in hits)


def test_search_caps_result_count():
    hits = search_tables(_tables(), "车间", limit=10)
    assert len(hits) <= 10


def test_placement_to_frontline_add():
    bonus = _bonus()
    placement = ManualPlacement(
        placement_id="t1",
        name="检索新人",
        eid="99900001",
        duty="操作工",
        workshop="101车间",
        hire_date=date(2024, 1, 1),
        target_sheet=SHEET_FRONTLINE,
        source_file="人员清单.xlsx",
        source_sheet="生产部",
        source_row=12,
    )
    frontline, supervisor = split_placements([placement], bonus)
    assert supervisor == []
    assert len(frontline) == 1
    assert frontline[0].action == ACTION_ADD
    assert frontline[0].category == CATEGORY_MANUAL
    assert frontline[0].workshop == "101车间"
    assert frontline[0].new_values["姓名"] == "检索新人"


def test_placement_to_frontline_update_when_already_there():
    bonus = _bonus()
    existing = next(p for p in bonus.frontline.values() if p.name == "陈玉慧")
    placement = ManualPlacement(
        placement_id="t2",
        name="陈玉慧",
        eid=existing.eid,
        duty="工艺员",
        workshop=existing.workshop,
        hire_date=None,
        target_sheet=SHEET_FRONTLINE,
        source_file="x.xlsx",
        source_sheet="s",
        source_row=3,
    )
    frontline, _ = split_placements([placement], bonus)
    assert frontline[0].action == ACTION_UPDATE
    assert frontline[0].frontline_row == existing.row


def test_placement_to_others_moves_off_frontline():
    bonus = _bonus()
    existing = next(p for p in bonus.frontline.values() if p.name == "陈玉慧")
    placement = ManualPlacement(
        placement_id="t3",
        name="陈玉慧",
        eid=existing.eid,
        duty="工艺员",
        workshop="101车间",
        hire_date=None,
        target_sheet=SHEET_OTHERS,
        source_file="x.xlsx",
        source_sheet="s",
        source_row=3,
    )
    frontline, supervisor = split_placements([placement], bonus)
    assert frontline[0].action == ACTION_MOVE
    assert supervisor == []


def test_placement_to_others_adds_when_absent():
    bonus = _bonus()
    placement = ManualPlacement(
        placement_id="t4",
        name="检索主任",
        eid="99900002",
        duty="车间副主任",
        workshop="101车间",
        hire_date=date(2020, 1, 1),
        target_sheet=SHEET_OTHERS,
        source_file="x.xlsx",
        source_sheet="s",
        source_row=9,
    )
    frontline, supervisor = split_placements([placement], bonus)
    assert frontline == []
    assert supervisor[0].action == ACTION_ADD
    assert supervisor[0].target_workshop == "101车间"


def test_placement_removes_from_supervisor_when_sent_to_frontline():
    bonus = _bonus()
    person = Person(
        name="检索主任",
        eid="99900002",
        duty="车间副主任",
        duty_raw="车间副主任",
        group="101车间",
        hire_date=None,
        row=4,
        workshop="101车间",
    )
    placement = ManualPlacement(
        placement_id="t5",
        name="检索主任",
        eid="99900002",
        duty="操作工",
        workshop="101车间",
        hire_date=None,
        target_sheet=SHEET_FRONTLINE,
        source_file="x.xlsx",
        source_sheet="s",
        source_row=9,
    )
    frontline, supervisor = split_placements([placement], bonus, others=[person])
    assert frontline[0].action == ACTION_ADD
    assert supervisor[0].action == ACTION_REMOVE
    assert supervisor[0].frontline_row == 4


def test_export_includes_manual_frontline_placement():
    bonus = _bonus()
    placement = ManualPlacement(
        placement_id="exp",
        name="检索投放甲",
        eid="88800001",
        duty="操作工",
        workshop="101车间",
        hire_date=date(2024, 6, 1),
        target_sheet=SHEET_FRONTLINE,
        source_file="人员清单.xlsx",
        source_sheet="生产部",
        source_row=20,
    )
    frontline, supervisor = split_placements([placement], bonus)
    adds, removes, updates, moves = split_by_action(frontline)
    data, _ = build_workbook(
        BONUS.read_bytes(),
        bonus,
        adds,
        removes,
        updates,
        moves,
        other_adds=supervisor,
        mode="apply",
    )
    wb = load_workbook(io.BytesIO(data))
    sheet = wb[SHEET_FRONTLINE]
    names = [str(sheet.cell(row, 3).value or "") for row in range(3, sheet.max_row + 1)]
    wb.close()
    assert "检索投放甲" in names
