import io
import zipfile

import openpyxl
import pytest

from tj4tools.bonus_export import build_workbook
from tj4tools.roster import ADD_CATEGORIES, CATEGORY_LEFT, CATEGORY_NEW

SHEET = "一线人员"


def _mapped(result, categories=None):
    """取有明确车间的人员，避免未映射项影响断言。"""
    items = [i for i in result.items if i.action == "add" and i.workshop]
    if categories:
        items = [i for i in items if i.category in categories]
    return items


@pytest.fixture(scope="module")
def marked(bonus_bytes, bonus, result):
    adds = _mapped(result)
    removes = [i for i in result.items if i.action == "remove"]
    updates = [i for i in result.items if i.action == "update"]
    data, summary = build_workbook(bonus_bytes, bonus, adds, removes, updates, mode="mark")
    return data, summary, adds, removes


def test_no_parts_lost_except_calc_chain(bonus_bytes, marked):
    data = marked[0]
    before = set(zipfile.ZipFile(io.BytesIO(bonus_bytes)).namelist())
    after = set(zipfile.ZipFile(io.BytesIO(data)).namelist())
    assert before - after == {"xl/calcChain.xml"}
    assert after - before == set()
    # 外部链接必须保留，否则副主任表的 VLOOKUP 全废
    assert "xl/externalLinks/externalLink1.xml" in after
    assert "xl/printerSettings/printerSettings1.bin" in after
    assert "xl/customProperty1.bin" in after
    assert "xl/comments1.xml" in after


def test_unmodified_parts_are_byte_identical(bonus_bytes, marked):
    source = zipfile.ZipFile(io.BytesIO(bonus_bytes))
    output = zipfile.ZipFile(io.BytesIO(marked[0]))
    changed = {"xl/worksheets/sheet2.xml", "xl/styles.xml", "xl/workbook.xml",
               "[Content_Types].xml", "xl/_rels/workbook.xml.rels"}
    for name in source.namelist():
        if name in changed or name == "xl/calcChain.xml":
            continue
        assert source.read(name) == output.read(name), name


def test_row_count_and_values(marked):
    data, summary, adds, removes = marked
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    sheet = workbook[SHEET]
    assert sheet.max_row == 676 + len(adds)
    assert summary.added == len(adds)
    assert summary.removed == len(removes)
    names = {
        (str(sheet.cell(row, 3).value or ""), str(sheet.cell(row, 4).value or ""))
        for row in range(3, sheet.max_row + 1)
    }
    for item in adds:
        assert (item.name, item.eid) in names, item


def test_inserted_rows_sit_at_bottom_of_their_duty_run(bonus_bytes, bonus, result):
    """新增助工要落在该车间助工那一段的最下面，而不是整个车间块的最下面。

    这里刻意只做新增：更新会改写「职务」列的值，会干扰"上一行是同职务"的判断。
    """
    adds = _mapped(result)
    data, _ = build_workbook(bonus_bytes, bonus, adds, [], [], mode="mark")
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    sheet = workbook[SHEET]
    blocks = {}
    for merge in sheet.merged_cells.ranges:
        if merge.min_col == 1 == merge.max_col and merge.min_row >= 3:
            label = sheet.cell(merge.min_row, 1).value
            if label:
                blocks[str(label).strip()] = (merge.min_row, merge.max_row)

    grouped = {}
    for item in adds:
        grouped.setdefault((item.workshop, item.duty), []).append(item)
    checked = 0
    for (workshop, duty), items in grouped.items():
        anchor, exact = bonus.anchor_for(workshop, duty)
        if not exact:
            continue
        start, end = blocks[workshop]
        rows = [
            row
            for row in range(start, end + 1)
            if (str(sheet.cell(row, 3).value or ""), str(sheet.cell(row, 4).value or ""))
            in {(i.name, i.eid) for i in items}
        ]
        assert len(rows) == len(items), (workshop, duty)
        # 必须连续，且紧接在该职务原有最后一行之后
        assert rows == list(range(rows[0], rows[0] + len(rows)))
        assert str(sheet.cell(rows[0] - 1, 2).value).strip() == duty, (workshop, duty)
        # 后面一行要么出块，要么是别的职务（说明我们确实停在这一段末尾）
        following = sheet.cell(rows[-1] + 1, 2).value
        if rows[-1] < end:
            assert str(following).strip() != duty, (workshop, duty)
        checked += 1
    assert checked >= 20


def test_duty_runs_stay_grouped_inside_each_block(bonus, marked):
    """插入后，每个车间块内同职务仍应聚在一起（段数不增加）。"""
    data, _, _, _ = marked
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    sheet = workbook[SHEET]
    for merge in sheet.merged_cells.ranges:
        if merge.min_col != 1 or merge.max_col != 1 or merge.min_row < 3:
            continue
        workshop = str(sheet.cell(merge.min_row, 1).value or "").strip()
        duties = [
            str(sheet.cell(row, 2).value or "").strip()
            for row in range(merge.min_row, merge.max_row + 1)
            if sheet.cell(row, 3).value
        ]
        runs = [d for index, d in enumerate(duties) if index == 0 or duties[index - 1] != d]
        original_runs = _run_count(
            [
                str(sheet.cell(row, 2).value or "").strip()
                for row in range(merge.min_row, merge.max_row + 1)
                if sheet.cell(row, 3).value
            ]
        )
        assert len(runs) == original_runs, (workshop, runs)


def _run_count(duties):
    return sum(1 for index, d in enumerate(duties) if index == 0 or duties[index - 1] != d)


def test_duty_run_count_does_not_grow(bonus_bytes, bonus, result):
    """插入后同职务不应被拆成更多段（更新写入的新职务除外，单独算）。"""
    adds = _mapped(result)
    data, _ = build_workbook(bonus_bytes, bonus, adds, [], [], mode="mark")
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    sheet = workbook[SHEET]
    for merge in sheet.merged_cells.ranges:
        if merge.min_col != 1 or merge.max_col != 1 or merge.min_row < 3:
            continue
        workshop = str(sheet.cell(merge.min_row, 1).value or "").strip()
        duties = [
            str(sheet.cell(row, 2).value or "").strip()
            for row in range(merge.min_row, merge.max_row + 1)
            if sheet.cell(row, 3).value
        ]
        _, start, end = bonus.block_of(workshop)
        before = _run_count(
            [p.duty for p in bonus.frontline.values() if start <= p.row <= end]
        )
        assert _run_count(duties) == before, (workshop, duties)


def test_inserted_row_formulas_reference_their_own_row(marked):
    data, _, adds, _ = marked
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    sheet = workbook[SHEET]
    checked = 0
    for row in range(3, sheet.max_row + 1):
        key = (str(sheet.cell(row, 3).value or ""), str(sheet.cell(row, 4).value or ""))
        if key not in {(i.name, i.eid) for i in adds}:
            continue
        assert sheet.cell(row, 6).value == f"=(K{row}+G{row}+H{row}+I{row}-J{row})"
        assert f"C{row}" in sheet.cell(row, 7).value
        assert f"L{row}:AP{row}" in sheet.cell(row, 11).value
        checked += 1
    assert checked == len(adds)


def test_new_rows_never_inherit_block_statistics_formulas(marked):
    """AQ/AR/AS 是按车间块统计的，只应留在原来的那些行上，不能被新行复制。"""
    data, _, adds, _ = marked
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    sheet = workbook[SHEET]
    add_keys = {(i.name, i.eid) for i in adds}
    for row in range(3, sheet.max_row + 1):
        key = (str(sheet.cell(row, 3).value or ""), str(sheet.cell(row, 4).value or ""))
        if key not in add_keys:
            continue
        for column in (43, 44, 45):
            assert sheet.cell(row, column).value is None, (row, column)


def test_broken_source_formula_is_not_propagated(bonus_bytes, bonus, result):
    """源文件 G342/H342/J342 本来就是 #REF!，新增行不能跟着坏掉。"""
    source = openpyxl.load_workbook(io.BytesIO(bonus_bytes), data_only=False)[SHEET]
    assert "#REF!" in source["G342"].value, "样本文件应当带着这个已知的坏公式"

    adds = [i for i in _mapped(result) if i.workshop == "5号楼" and i.duty == "助工"]
    assert adds
    data, summary = build_workbook(bonus_bytes, bonus, adds, [], [], mode="mark")
    assert any("#REF!" in warning for warning in summary.warnings)
    sheet = openpyxl.load_workbook(io.BytesIO(data), data_only=False)[SHEET]
    add_keys = {(i.name, i.eid) for i in adds}
    checked = 0
    for row in range(3, sheet.max_row + 1):
        key = (str(sheet.cell(row, 3).value or ""), str(sheet.cell(row, 4).value or ""))
        if key not in add_keys:
            continue
        for column in (7, 8, 10):
            formula = sheet.cell(row, column).value
            assert "#REF!" not in formula, (row, column, formula)
            assert f"C{row}" in formula
        checked += 1
    assert checked == len(adds)


def test_existing_row_formulas_follow_their_new_row(marked):
    data, _, _, _ = marked
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    sheet = workbook[SHEET]
    for row in range(3, sheet.max_row):
        formula = sheet.cell(row, 6).value
        if isinstance(formula, str) and formula.startswith("="):
            assert formula == f"=(K{row}+G{row}+H{row}+I{row}-J{row})", row


def test_workshop_blocks_grow_by_their_own_additions(bonus, marked):
    data, _, adds, _ = marked
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    sheet = workbook[SHEET]
    blocks = sorted(
        (merge.min_row, merge.max_row)
        for merge in sheet.merged_cells.ranges
        if merge.min_col == 1 == merge.max_col and merge.min_row >= 3
    )
    assert len(blocks) == len(bonus.blocks)
    per_workshop = {}
    for item in adds:
        per_workshop[item.workshop] = per_workshop.get(item.workshop, 0) + 1
    for (start, end), (workshop, old_start, old_end) in zip(blocks, bonus.blocks):
        assert str(sheet.cell(start, 1).value).strip() == workshop
        expected = (old_end - old_start + 1) + per_workshop.get(workshop, 0)
        assert end - start + 1 == expected, workshop


def test_block_statistics_ranges_cover_new_rows(bonus, marked):
    """AQ/AR/AS 的 a/b/c 分布按块统计，区间必须把落进该区间的新增人员包进来。

    注意原表 AQ525 的区间是 $L$523:$AP$630，横跨 DCS/计算机化/培训 三个车间块，
    所以断言写成"按原区间覆盖的车间累加新增人数"，而不是假设区间等于单个块。
    """
    import re

    data, _, adds, _ = marked
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    sheet = workbook[SHEET]

    per_workshop = {}
    for item in adds:
        per_workshop[item.workshop] = per_workshop.get(item.workshop, 0) + 1
    # 原表每个车间块末行收到的新增人数
    added_at = {end: per_workshop.get(name, 0) for name, _, end in bonus.blocks}
    ends = sorted(added_at)

    def shift(row, inclusive):
        return row + sum(added_at[e] for e in ends if (e <= row if inclusive else e < row))

    original_ranges = [
        (3, 73), (74, 164), (165, 226), (227, 298), (299, 349), (350, 522),
        (523, 630), (624, 630), (631, 643), (644, 656), (657, 675),
    ]
    expected = {(shift(a, False), shift(b, True)) for a, b in original_ranges}

    pattern = re.compile(r"\$?L\$?(\d+):\$?AP\$?(\d+)")
    found = 0
    for row in range(1, sheet.max_row + 1):
        for column in (43, 44, 45):
            formula = sheet.cell(row, column).value
            if not isinstance(formula, str) or "COUNTIF" not in formula:
                continue
            for start, end in pattern.findall(formula):
                assert (int(start), int(end)) in expected, (row, column, formula)
                found += 1
    assert found == len(original_ranges) * 3


def test_total_row_sum_extends(marked):
    data, _, adds, _ = marked
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    sheet = workbook[SHEET]
    last = sheet.max_row
    assert sheet.cell(last, 6).value == f"=SUM(F3:J{last - 1})"


def test_autofilter_and_defined_names_updated(marked):
    data, _, adds, _ = marked
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    sheet = workbook[SHEET]
    assert sheet.auto_filter.ref == f"A1:AU{sheet.max_row}"
    text = zipfile.ZipFile(io.BytesIO(data)).read("xl/workbook.xml").decode("utf-8")
    assert f"一线人员!$A$1:$AU${sheet.max_row}" in text
    assert f"一线人员!$A$1:$F${sheet.max_row - 1}" in text
    assert 'fullCalcOnLoad="1"' in text


def test_new_rows_are_green_and_removed_rows_red(marked):
    data, _, adds, removes = marked
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    sheet = workbook[SHEET]
    add_keys = {(i.name, i.eid) for i in adds}
    remove_keys = {(i.name, i.eid) for i in removes}
    greens = reds = 0
    for row in range(3, sheet.max_row + 1):
        name_cell = sheet.cell(row, 3)
        eid_cell = sheet.cell(row, 4)
        key = (str(name_cell.value or ""), str(eid_cell.value or ""))
        colors = {name_cell.fill.fgColor.rgb, eid_cell.fill.fgColor.rgb}
        if key in add_keys:
            assert colors == {"FF92D050"}, key
            greens += 1
        elif key in remove_keys or _strip(key) in remove_keys:
            assert colors == {"FFFF0000"}, key
            reds += 1
    assert greens == len(adds)
    assert reds == len(removes)


def test_apply_mode_marks_new_rows_with_red_font(bonus_bytes, bonus, result):
    adds = _mapped(result)[:12]
    data, _ = build_workbook(bonus_bytes, bonus, adds, [], [], mode="apply")
    sheet = openpyxl.load_workbook(io.BytesIO(data))[SHEET]
    add_keys = {(i.name, i.eid) for i in adds}
    checked = 0
    for row in range(3, sheet.max_row + 1):
        key = (str(sheet.cell(row, 3).value or ""), str(sheet.cell(row, 4).value or ""))
        if key not in add_keys:
            continue
        reference = sheet.cell(4, 3)
        for column in (2, 3, 4, 5):
            cell = sheet.cell(row, column)
            assert cell.font.color is not None and cell.font.color.rgb == "FFFF0000", (row, column)
            # 只改颜色，字号字体保持不变
            assert cell.font.name == reference.font.name
            assert cell.font.sz == reference.font.sz
            assert cell.fill.fgColor.rgb != "FF92D050"
        checked += 1
    assert checked == len(adds)
    # 原有行的字体颜色不能被动到
    assert sheet.cell(4, 3).font.color is None or sheet.cell(4, 3).font.color.rgb != "FFFF0000"


def test_interns_get_a_yellow_duty_cell(bonus_bytes, roster_bytes, bonus):
    from tj4tools.roster import parse_roster, reconcile

    roster = parse_roster(
        roster_bytes, "TJ4生产部&生产设备部人员清单，07-31-2026.xlsx", include_interns=True
    )
    analysis = reconcile(roster, bonus)
    interns = [i for i in analysis.items if i.action == "add" and i.workshop and i.is_intern]
    assert len(interns) >= 30, "纳入实习生后应当有一批实习生待新增"
    for mode in ("mark", "apply"):
        data, summary = build_workbook(bonus_bytes, bonus, interns, [], [], mode=mode)
        assert summary.interns == len(interns)
        sheet = openpyxl.load_workbook(io.BytesIO(data))[SHEET]
        keys = {(i.name, i.eid) for i in interns}
        found = 0
        for row in range(3, sheet.max_row + 1):
            key = (str(sheet.cell(row, 3).value or ""), str(sheet.cell(row, 4).value or ""))
            if key not in keys:
                continue
            assert sheet.cell(row, 2).fill.fgColor.rgb == "FFFFFF00", (mode, row)
            found += 1
        assert found == len(interns), mode


def test_updates_rewrite_the_row_with_roster_values(bonus_bytes, bonus, result):
    updates = result.by_action("update")
    assert len(updates) == 1
    renamed = next(i for i in updates if i.key == ("曹静旺", "ALS14679"))

    data, summary = build_workbook(bonus_bytes, bonus, [], [], updates, mode="mark")
    assert summary.updated == 1
    sheet = openpyxl.load_workbook(io.BytesIO(data))[SHEET]
    # 没有插入也没有删除，行号保持原样
    assert sheet.max_row == 676
    assert sheet.cell(renamed.frontline_row, 3).value == "曹睿晟"
    assert sheet.cell(renamed.frontline_row, 3).fill.fgColor.rgb == "FFFFC000"
    # 没有差异的列不能被着色
    assert sheet.cell(renamed.frontline_row, 4).fill.fgColor.rgb != "FFFFC000"

    plain, _ = build_workbook(bonus_bytes, bonus, [], [], updates, mode="apply")
    sheet = openpyxl.load_workbook(io.BytesIO(plain))[SHEET]
    assert sheet.cell(renamed.frontline_row, 3).value == "曹睿晟"
    assert sheet.cell(renamed.frontline_row, 3).fill.fgColor.rgb != "FFFFC000"


def test_update_keeps_the_row_formulas_intact(bonus_bytes, bonus, result):
    updates = result.by_action("update")
    data, _ = build_workbook(bonus_bytes, bonus, [], [], updates, mode="apply")
    sheet = openpyxl.load_workbook(io.BytesIO(data), data_only=False)[SHEET]
    for item in updates:
        row = item.frontline_row
        assert sheet.cell(row, 6).value == f"=(K{row}+G{row}+H{row}+I{row}-J{row})"
        assert sheet.cell(row, 5).is_date or sheet.cell(row, 5).value is None


OTHERS = "副主任&工艺组长及其他"


def test_moves_leave_the_frontline_and_land_in_the_other_sheet(bonus_bytes, bonus, result):
    moves = result.by_action("move")
    assert len(moves) == 15
    data, summary = build_workbook(bonus_bytes, bonus, [], [], [], moves, mode="apply")
    assert summary.moved == 15
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    frontline = workbook[SHEET]
    others = workbook[OTHERS]
    assert frontline.max_row == 676 - 15
    assert others.max_row == 206 + 15

    gone = {
        (str(frontline.cell(r, 3).value or ""), str(frontline.cell(r, 4).value or ""))
        for r in range(3, frontline.max_row + 1)
    }
    landed = {
        (str(others.cell(r, 3).value or ""), str(others.cell(r, 4).value or ""))
        for r in range(2, others.max_row + 1)
    }
    for item in moves:
        assert (item.name, item.eid) not in gone, item.name
        assert (str(item.new_values["姓名"]), str(item.new_values["员工编号"])) in landed, item.name


def test_moved_rows_carry_workshop_duty_and_formulas(bonus_bytes, bonus, result):
    moves = result.by_action("move")
    data, _ = build_workbook(bonus_bytes, bonus, [], [], [], moves, mode="apply")
    sheet = openpyxl.load_workbook(io.BytesIO(data), data_only=False)[OTHERS]
    by_name = {}
    for row in range(2, sheet.max_row + 1):
        name = str(sheet.cell(row, 3).value or "")
        if name:
            by_name.setdefault(name, row)
    for item in moves:
        row = by_name[str(item.new_values["姓名"])]
        # 这张表的车间列不合并，每行都要写车间名
        assert str(sheet.cell(row, 1).value or "").strip() == item.target_workshop, item.name
        assert str(sheet.cell(row, 2).value or "").strip() == item.new_values["职务"], item.name
        assert sheet.cell(row, 5).is_date
        # 该表的每行公式（F/G/H/J/L/R）要指向自己这一行
        assert sheet.cell(row, 6).value == f"=(L{row}*K{row}+G{row}+H{row}+I{row}-J{row})"
        assert f"D{row}" in sheet.cell(row, 18).value
        assert "[1]生产部" in sheet.cell(row, 18).value


def test_move_shifts_the_lookup_table_reference(bonus_bytes, bonus, result):
    """副主任表 L 列的 HLOOKUP 引用第 143/144 行的档位参照表，插行后必须跟着顺延。"""
    source = openpyxl.load_workbook(io.BytesIO(bonus_bytes), data_only=False)[OTHERS]
    assert source["L2"].value == "=HLOOKUP(M2,$A$143:$H$144,2)"
    moves = result.by_action("move")
    data, _ = build_workbook(bonus_bytes, bonus, [], [], [], moves, mode="apply")
    sheet = openpyxl.load_workbook(io.BytesIO(data), data_only=False)[OTHERS]
    assert sheet["L2"].value == f"=HLOOKUP(M2,$A${143 + 15}:$H${144 + 15},2)"
    assert sheet.cell(143 + 15, 1).value == source.cell(143, 1).value


def test_moves_go_to_the_right_workshop_block(bonus_bytes, bonus, result):
    moves = result.by_action("move")
    data, summary = build_workbook(bonus_bytes, bonus, [], [], [], moves, mode="apply")
    sheet = openpyxl.load_workbook(io.BytesIO(data))[OTHERS]
    rows = []
    for row in range(2, sheet.max_row + 1):
        name = str(sheet.cell(row, 3).value or "")
        if not name:
            break
        rows.append((row, str(sheet.cell(row, 1).value or "").strip(), name))
    # 车间仍然成块，不能被打散
    seen: list[str] = []
    for _, workshop, _ in rows:
        if not seen or seen[-1] != workshop:
            assert workshop not in seen, f"车间 {workshop} 被拆成了多段"
            seen.append(workshop)
    # 副主任表原本没有的车间会追加在最后
    assert set(summary.new_other_workshops) == {"清洗组", "DCS", "外围/罐区/泵房"}
    assert seen[-3:] == ["DCS", "外围/罐区/泵房", "清洗组"]


def test_marked_mode_flags_movers_red_in_the_frontline(bonus_bytes, bonus, result):
    moves = result.by_action("move")
    data, _ = build_workbook(bonus_bytes, bonus, [], [], [], moves, mode="mark")
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    frontline = workbook[SHEET]
    others = workbook[OTHERS]
    assert frontline.max_row == 676, "标记版不删行"
    for item in moves:
        assert frontline.cell(item.frontline_row, 3).fill.fgColor.rgb == "FFFF0000"
        assert frontline.cell(item.frontline_row, 4).fill.fgColor.rgb == "FFFF0000"
    landed = [
        row
        for row in range(2, others.max_row + 1)
        if str(others.cell(row, 3).value or "") in {str(i.new_values["姓名"]) for i in moves}
    ]
    assert len(landed) == len(moves)
    for row in landed:
        assert others.cell(row, 3).fill.fgColor.rgb == "FF92D050"


def test_other_sheets_still_untouched_when_moving(bonus_bytes, bonus, result):
    """改了两张子表，其余六张必须一字不动。"""
    moves = result.by_action("move")
    data, _ = build_workbook(bonus_bytes, bonus, [], [], [], moves, mode="apply")
    before = openpyxl.load_workbook(io.BytesIO(bonus_bytes), data_only=False)
    after = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    assert before.sheetnames == after.sheetnames
    for name in before.sheetnames:
        if name in (SHEET, OTHERS):
            continue
        source, target = before[name], after[name]
        assert source.max_row == target.max_row, name
        for row in source.iter_rows():
            for cell in row:
                assert target[cell.coordinate].value == cell.value, (name, cell.coordinate)


def _strip(key):
    from tj4tools.normalize import norm_eid, norm_name

    return (norm_name(key[0]), norm_eid(key[1]))


def test_formatting_of_inserted_rows_matches_template(marked):
    data, _, adds, _ = marked
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    sheet = workbook[SHEET]
    reference = sheet.cell(4, 2)
    add_keys = {(i.name, i.eid) for i in adds}
    for row in range(3, sheet.max_row + 1):
        key = (str(sheet.cell(row, 3).value or ""), str(sheet.cell(row, 4).value or ""))
        if key not in add_keys:
            continue
        cell = sheet.cell(row, 2)
        assert cell.font.name == reference.font.name
        assert cell.font.sz == reference.font.sz
        assert cell.alignment.horizontal == reference.alignment.horizontal
        assert sheet.row_dimensions[row].height == sheet.row_dimensions[4].height
        # 入职日期应写成真正的日期而不是文本
        assert sheet.cell(row, 5).is_date
        assert sheet.cell(row, 5).number_format == sheet.cell(4, 5).number_format


def test_other_sheets_untouched(bonus_bytes, marked):
    before = openpyxl.load_workbook(io.BytesIO(bonus_bytes), data_only=False)
    after = openpyxl.load_workbook(io.BytesIO(marked[0]), data_only=False)
    assert before.sheetnames == after.sheetnames
    for name in before.sheetnames:
        if name == SHEET:
            continue
        source, target = before[name], after[name]
        assert source.max_row == target.max_row, name
        assert source.sheet_state == target.sheet_state, name
        assert source.auto_filter.ref == target.auto_filter.ref, name
        for row in source.iter_rows():
            for cell in row:
                assert target[cell.coordinate].value == cell.value, (name, cell.coordinate)


def test_apply_mode_deletes_and_inserts_without_color(bonus_bytes, bonus, result):
    adds = _mapped(result, ADD_CATEGORIES)[:5]
    removes = [i for i in result.items if i.category == CATEGORY_LEFT][:4]
    data, summary = build_workbook(bonus_bytes, bonus, adds, removes, mode="apply")
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    sheet = workbook[SHEET]
    assert summary.added == len(adds)
    assert summary.removed == len(removes)
    assert sheet.max_row == 676 + len(adds) - len(removes)
    present = {
        (str(sheet.cell(row, 3).value or ""), str(sheet.cell(row, 4).value or ""))
        for row in range(3, sheet.max_row + 1)
    }
    for item in removes:
        assert (item.name, item.eid) not in present
    for item in adds:
        assert (item.name, item.eid) in present
        row = next(r for r in range(3, sheet.max_row + 1)
                   if str(sheet.cell(r, 3).value or "") == item.name)
        assert sheet.cell(row, 3).fill.fgColor.rgb not in ("FF92D050", "FFFF0000")
        assert sheet.cell(row, 4).fill.fgColor.rgb not in ("FF92D050", "FFFF0000")


def test_apply_mode_keeps_formulas_consistent(bonus_bytes, bonus, result):
    adds = _mapped(result, [CATEGORY_NEW])[:6]
    removes = [i for i in result.items if i.category == CATEGORY_LEFT][:6]
    data, _ = build_workbook(bonus_bytes, bonus, adds, removes, mode="apply")
    workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    sheet = workbook[SHEET]
    for row in range(3, sheet.max_row):
        formula = sheet.cell(row, 6).value
        if isinstance(formula, str) and formula.startswith("="):
            assert formula == f"=(K{row}+G{row}+H{row}+I{row}-J{row})", row
    merged = [m for m in sheet.merged_cells.ranges if m.min_col == 1 == m.max_col and m.min_row >= 3]
    merged.sort(key=lambda m: m.min_row)
    assert merged[0].min_row == 3
    for previous, current in zip(merged, merged[1:]):
        assert current.min_row == previous.max_row + 1


def test_new_workshop_block_is_created(bonus_bytes, bonus, result):
    items = [i for i in result.items if i.action == "add"][:3]
    for item in items:
        item.workshop = "13号楼"
    data, summary = build_workbook(bonus_bytes, bonus, items, [], mode="mark")
    assert summary.new_blocks == ["13号楼"]
    assert any("新建分组块" in w for w in summary.warnings)
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    sheet = workbook[SHEET]
    # 新块紧跟在最后一个数据行之后，原本的合计行被顺延
    assert sheet.cell(676, 1).value == "13号楼"
    assert any(str(m) == "A676:A678" for m in sheet.merged_cells.ranges)
    # 原最后一个车间块不能被新块吞掉
    assert any(str(m) == "A657:A675" for m in sheet.merged_cells.ranges)
    assert sheet.cell(679, 6).value == "=SUM(F3:J678)"


def test_items_without_workshop_are_skipped(bonus_bytes, bonus, result):
    orphans = [i for i in result.items if i.action == "add" and not i.workshop]
    if not orphans:
        pytest.skip("样本里没有未映射人员")
    data, summary = build_workbook(bonus_bytes, bonus, orphans[:3], [], mode="mark")
    assert summary.added == 0
    assert len(summary.skipped) == 3
    workbook = openpyxl.load_workbook(io.BytesIO(data))
    assert workbook[SHEET].max_row == 676
